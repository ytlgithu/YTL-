# -*- coding: utf-8 -*-
import os
os.chdir(r"D:\桌面文件\ytl\QClawProject\YTL仓博系统")
import sys
sys.path.insert(0, r"D:\桌面文件\ytl\QClawProject\YTL仓博系统")

import openpyxl

fpath = 'uploads/repo_4/三轴点胶IO表.xlsx'
wb = openpyxl.load_workbook(fpath, data_only=False)
ws = wb.active

print("=== 检查第一个单元格 (行1列1) ===")
cell = ws.cell(1, 1)
print(f"值: {cell.value}")
print(f"fill: {cell.fill}")
print(f"fill type: {type(cell.fill)}")
if cell.fill:
    print(f"fill_type: {cell.fill.fill_type}")
    print(f"fgColor: {cell.fill.fgColor}")
    if cell.fill.fgColor:
        print(f"  fgColor.rgb: {cell.fill.fgColor.rgb}")
        print(f"  fgColor.type: {cell.fill.fgColor.type}")

print("\n=== 检查第二行第一列 (输入) ===")
cell2 = ws.cell(2, 1)
print(f"值: {cell2.value}")
print(f"fill_type: {cell2.fill.fill_type if cell2.fill else None}")
if cell2.fill and cell2.fill.fgColor:
    print(f"fgColor.rgb: {cell2.fill.fgColor.rgb}")

print("\n=== 检查第三行第一列 (1) ===")
cell3 = ws.cell(3, 1)
print(f"值: {cell3.value}")
if cell3.fill:
    print(f"fill_type: {cell3.fill.fill_type}")
    if cell3.fill.fgColor:
        print(f"fgColor.rgb: {cell3.fill.fgColor.rgb}")

wb.close()